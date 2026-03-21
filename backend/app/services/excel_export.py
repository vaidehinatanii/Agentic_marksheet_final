"""Excel export service with conditional formatting."""
import io
import logging
from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import MarksheetRecord, SubjectStatus

logger = logging.getLogger(__name__)

# Cell styles
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)


def create_excel(records: List[MarksheetRecord]) -> bytes:
    """
    Create Excel file from records with formatting.
    Includes all 6 subjects (English, Math, Science, Social Science, Second Language, Additional).
    Returns file bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Marksheet Results"

    # Define columns - Basic info + 6 subjects (each with name, obtained, max, status) + computed
    columns = [
        "Student Name",
        "Roll No",
        "Exam Session",
        "School",
        "Result Status",
        # Subject 1: English
        "Subj 1 Name",
        "Subj 1 Obtained",
        "Subj 1 Max",
        "Subj 1 Status",
        # Subject 2: Mathematics
        "Subj 2 Name",
        "Subj 2 Obtained",
        "Subj 2 Max",
        "Subj 2 Status",
        # Subject 3: Science
        "Subj 3 Name",
        "Subj 3 Obtained",
        "Subj 3 Max",
        "Subj 3 Status",
        # Subject 4: Social Science
        "Subj 4 Name",
        "Subj 4 Obtained",
        "Subj 4 Max",
        "Subj 4 Status",
        # Subject 5: Second Language
        "Subj 5 Name",
        "Subj 5 Obtained",
        "Subj 5 Max",
        "Subj 5 Status",
        # Subject 6: Additional
        "Subj 6 Name",
        "Subj 6 Obtained",
        "Subj 6 Max",
        "Subj 6 Status",
        # Computed fields
        "Best 5 %",
        "Core % (Eng, Math, Sci, SS)",
        "Needs Review",
        "Review Reasons"
    ]

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGNMENT
        cell.border = BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set column widths
    column_widths = {
        "A": 25,  # Student Name
        "B": 15,  # Roll No
        "C": 15,  # Exam Session
        "D": 30,  # School
        "E": 12,  # Result Status
        # Subjects (4 columns each: name, obtained, max, status)
        "F": 18, "G": 12, "H": 8, "I": 12,  # Subj 1
        "J": 18, "K": 12, "L": 8, "M": 12,  # Subj 2
        "N": 18, "O": 12, "P": 8, "Q": 12,  # Subj 3
        "R": 18, "S": 12, "T": 8, "U": 12,  # Subj 4
        "V": 18, "W": 12, "X": 8, "Y": 12,  # Subj 5
        "Z": 18, "AA": 12, "AB": 8, "AC": 12,  # Subj 6
        "AD": 12,  # Best 5 %
        "AE": 20,  # Core %
        "AF": 12,  # Needs Review
        "AG": 40   # Review Reasons
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Write data rows
    for row_idx, record in enumerate(records, 2):
        # Organize subjects by position (first 6 subjects)
        subjects_list = record.subjects[:6] if len(record.subjects) >= 6 else record.subjects + [None] * (6 - len(record.subjects))

        # Basic info
        row_data = [
            record.student_name or "",
            record.roll_no or "",
            record.exam_session or "",
            record.school or "",
            record.result_status.value if record.result_status else "",
        ]

        # Add subject data (6 subjects)
        for subject in subjects_list:
            if subject and subject.normalized_name != "EXTRA":
                row_data.extend([
                    subject.normalized_name,
                    subject.obtained_marks if subject.obtained_marks is not None else "",
                    subject.max_marks if subject.max_marks is not None else "",
                    subject.status.value if subject.status else ""
                ])
            else:
                row_data.extend(["", "", "", ""])

        # Computed fields
        row_data.extend([
            record.overall_percent if record.overall_percent is not None else "",
            record.pcm_percent if record.pcm_percent is not None else "",  # Core %
            "Yes" if record.needs_review else "No",
            "; ".join(record.review_reasons) if record.review_reasons else ""
        ])

        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = ALIGNMENT
            cell.border = BORDER

            # Apply formatting for problematic cells
            if record.needs_review:
                col_letter = get_column_letter(col_idx)

                # Highlight subject cells with special status or missing values
                # Subject columns start at column 6 (F), each subject has 4 columns
                if col_idx >= 6 and col_idx <= 29:  # Subject columns
                    subject_col = (col_idx - 6) % 4
                    subject_idx = (col_idx - 6) // 4

                    if subject_idx < len(subjects_list):
                        subj = subjects_list[subject_idx]
                        if subj and subj.normalized_name != "EXTRA":
                            # Status column (index 3 in the 4-column group)
                            if subject_col == 3 and subj.status != SubjectStatus.OK:
                                cell.fill = ERROR_FILL
                            # Obtained column (index 1)
                            elif subject_col == 1 and subj.obtained_marks is None:
                                cell.fill = WARNING_FILL
                            # Max column (index 2)
                            elif subject_col == 2 and subj.max_marks is None:
                                cell.fill = WARNING_FILL

                # Highlight review columns
                if col_idx in [32, 33]:  # Needs Review, Review Reasons columns
                    cell.fill = ERROR_FILL

    logger.info("Excel export created with %d records", len(records))

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
